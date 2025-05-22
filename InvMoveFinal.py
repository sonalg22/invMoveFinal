#!/usr/bin/env python
# coding: utf-8

# In[174]:


get_ipython().system('pip install google-auth google-auth-oauthlib')
get_ipython().system('pip install google-api-python-client')
get_ipython().system('pip install psycopg2-binary')


# In[1]:


from flask import Flask, render_template, redirect, url_for
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from dotenv import load_dotenv
import googleapiclient.discovery
import os
import openpyxl
import re
import pandas as pd
import numpy as np
import psycopg2
import sqlalchemy
import pandas.io.sql as psql
import shutil
import requests
import json


# In[2]:


# read and load pick ticket file and inventory file
def load_data(pick_ticket_file, inventory_file):
    pick_data = pd.read_csv(pick_ticket_file, encoding='latin1') 
    inventory_data = pd.read_csv(inventory_file, encoding='latin1')
    return pick_data, inventory_data


# function to compare row counts for pick ticket data
def check_row_count(initial, current, step_name):
    if initial < current:
        raise ValueError(f"Row count mismatch at {step_name}! Expected: {expected}, Found: {current}")


# data transofations for pick ticket
def process_pick_data(pick_data, ab_or_gf): 
    
    #pick ticket data frame
    pick_data['staged_for_production'] = ""
    pick_data['used_for_production'] = ""
    pick_data['qty_returned_to_inventory'] = ""
    pick_data = pick_data.dropna(axis=1, how='all')


    pick_data = pick_data.rename(columns={
        'PICKITEMPARTNUM': 'part_number',
        'PICKITEMPARTDESC': 'description',
        'PICKITEMQTYVUOM': 'pick_qty',
        'VendorUOM': 'vendor_uom',
        'CasesNeeded': 'cases_needed',
        'CasesNeeded_v2': 'cases_rounded',
        'MONUM': 'mo_number',
        'PICKITEMUOM': 'pick_uom'
    })



    # drop uneeded columns
    pick_data = pick_data.drop(['PICKDATESCHEDULED', 'PICKITEMSTATUS', 'PICKITEMQTY'], axis=1)
    grouped_parts = pick_data.groupby('part_number')[['pick_qty', 'cases_needed', 'cases_rounded']].sum().reset_index() 

    # drop uneeded columns
    pick_data =  pick_data.drop(['pick_qty', 'cases_needed', 'cases_rounded'], axis=1)
    pick_data = pick_data.merge(grouped_parts, on = 'part_number', how = 'left')
    
    # reindex columns to move the last column to the 5th position
    cols = pick_data.columns.tolist()
    last_col = cols.pop()
    cols.insert(4, last_col)
    pick_data = pick_data.reindex(columns=cols)

    return pick_data





# data transofations for inventory sheet
def process_inventory_data(inventory_data, ab_or_gf):

    # filter inventory data
    inventory_filtered = inventory_data[inventory_data['Location'].str.contains(ab_or_gf)]
    #inventory_sheet = inventory_filtered[['PartNumber', 'Location', 'Qty', 'UOM', 'Tracking-Lot Number', 'Tracking-Expiration Date']]
    inventory_sheet = inventory_filtered[['PartNumber', 'PartDescription', 'Location', 'Qty', 'UOM', 'Tracking-Lot Number', 'Tracking-Expiration Date']]

    
    inventory_sheet = inventory_sheet.rename(columns={
        'PartNumber': 'part_number',
        'Location': 'beginning_location',
        'Qty': 'on_hand',
        'UOM': 'uom',
        'Tracking-Lot Number': 'lot_number',
        'Tracking-Expiration Date': 'expiration_date'
    })
    
    # new columns
    inventory_sheet['end_location'] = f"{ab_or_gf}_Meal Kit-Picking" 
    inventory_sheet['note'] = ""

    return inventory_sheet


# merges pick data and new inventory sheet
def merge_data(pick_data, inventory_sheet):
    master_data = pick_data.merge(inventory_sheet, on='part_number', how='inner') 

    part_counts = master_data['part_number'].value_counts().reset_index()
    part_counts.columns = ['part_number', 'part_count']

    master_data = master_data.merge(part_counts, on='part_number', how='left')

    #print(master_data[['part_number','part_count']].head())

    return master_data



# format, clean, and perform final transformations on the merged data.
def format_and_clean_master_data(master_data):

    # basic transformations
    master_data['short_quantity'] = ""
    master_data = master_data[['mo_number', 'beginning_location', 'lot_number', 'expiration_date', 'part_number', 'description', 'on_hand', 'uom', 'pick_qty',
                               'vendor_uom', 'cases_needed', 'cases_rounded', 'staged_for_production',
                               'used_for_production', 'qty_returned_to_inventory', 'part_count','short_quantity']]

    # convert 'vendor_uom' column to numeric where applicable #changed
    master_data.loc[master_data['vendor_uom'] == 'ea', 'vendor_uom'] = 1
    master_data.loc[:, 'vendor_uom'] = master_data['vendor_uom'].copy().apply(drop_letters_after_number)


    # exclude parts that are either in the Dandee location OR are already in a picking location
    master_data = master_data[~master_data['beginning_location'].str.contains('Dandee|Picking|Prepared|Receiving')] 
    master_data = master_data[~master_data['part_number'].str.startswith(('M-P', 'MK ', 'M-TA71103'))] 

    # replace NaN values in specific columns
    columns_to_replace_nan = ['beginning_location', 'lot_number', 'on_hand', 'cases_needed', 'cases_rounded']
    for column in columns_to_replace_nan:
        master_data[column] = master_data[column].fillna(0)

    # Set "move" column to FALSE (unchecked by default)
    master_data['move'] = ""
    master_data['substitution'] = ""
    
    # resort specific values
    master_data = master_data.sort_values(['part_number', 'lot_number', 'expiration_date'], ascending=True)
    # break this apart innto 3 lines of code
    
    # drop duplicated values
    master_data = master_data.drop_duplicates()

    # format numerical fields
    for col in ['vendor_uom', 'pick_qty', 'cases_needed', 'cases_rounded', 'on_hand']:
        master_data[col] = pd.to_numeric(master_data[col], errors='coerce')

    # round numerical fields
    master_data['pick_qty'] = round(master_data['pick_qty'], 2)
    master_data['cases_on_hand'] = round(master_data['on_hand'] / master_data['vendor_uom'], 2)
    master_data['part_count']= master_data.groupby('part_number').cumcount() + 1

    return master_data

# helper to clean up the vendor uom data column
def drop_letters_after_number(input_string):
    if isinstance(input_string, str):
        match = re.search(r'\d+(\.\d+)?', input_string)
        return float(match.group()) if match else 0
    return input_string





# function for adding checkboxes to move column
def addCheckboxesAndSyncMove(sheet_name):
    
    # get the sheetId from the sheet name
    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets = spreadsheet.get('sheets', [])
    
    # find sheetId for the sheet_name
    sheet_id = None
    for sheet in sheets:
        if sheet['properties']['title'] == sheet_name:
            sheet_id = sheet['properties']['sheetId']
            print(f"Found sheetId for '{sheet_name}': {sheet_id}")
            break
    
    if sheet_id is None:
        print(f"Error: Sheet with name '{sheet_name}' not found.")
        return

    # set the column number for 'Move' and 'Used for Production'
    header_row = 1
    move_col = None
    used_for_prod_col = None

    # get the sheet data to find the columns for 'Move' and 'Used for Production'
    range_to_check = f"{sheet_name}!A1:Z1"
    sheet_values = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=range_to_check
    ).execute()

    headers = sheet_values.get('values', [])[0]
    
    try:
        move_col = headers.index('Move') + 1  # 1-based index for columns
        used_for_prod_col = headers.index('Used for Production') + 1
    except ValueError:
        print(f"Error: Columns 'Move' or 'Used for Production' not found in the header row.")
        return

    # get all values in the sheet to determine the last used row
    range_data = f"{sheet_name}!A1:Z"
    sheet_values = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=range_data,
        majorDimension="ROWS"
    ).execute()
    
    # find the last row with actual data
    rows = sheet_values.get('values', [])
    last_used_row = len(rows)  # Number of rows with data
    
    # make sure checkboxes aren't infinite
    if last_used_row < 2:
        print("No data found in the sheet beyond headers.")
        last_used_row = 2  
    
    start_row_index = 1  
    end_row_index = last_used_row  

    # apply checkboxes to the 'Move' column -- requests starting
    requests = []

    # hardcoded to clear the entire UFP column
    requests.append({
        'updateCells': {
            'range': {
                'sheetId': sheet_id,
                'startRowIndex': start_row_index,  # Below the header
                'endRowIndex': end_row_index,  # Up to the last used row
                'startColumnIndex': used_for_prod_col - 1,  # 0-based index
                'endColumnIndex': used_for_prod_col
            },
            'rows': [
                {
                    'values': [
                        {
                            'userEnteredValue': None  # Makes the cell empty
                        }
                    ]
                } for _ in range(end_row_index - start_row_index)
            ],
            'fields': 'userEnteredValue'
        }
    })


    # insert checkboxes in the 'Move' column
    requests.append({
        'updateCells': {
            'range': {
                'sheetId': sheet_id,
                'startRowIndex': start_row_index,  # start at row 2 to avoid the header
                'endRowIndex': end_row_index,  
                'startColumnIndex': move_col - 1,  # converting to 0-based index
                'endColumnIndex': move_col
            },
            'rows': [
                {
                    'values': [
                        {
                            'userEnteredValue': {'boolValue': False},  
                            'dataValidation': {
                                'condition': {
                                    'type': 'BOOLEAN',
                                    'values': [
                                        {'userEnteredValue': 'TRUE'},
                                        {'userEnteredValue': 'FALSE'}
                                    ]
                                },
                                'showCustomUi': True
                            }
                        }
                    ]
                } for _ in range(end_row_index - start_row_index)  
            ],
            'fields': 'dataValidation'
        }
    })

    # get data to update checkboxes in the 'Move' column based on UFP column
    data = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{sheet_name}!A2:Z500"
    ).execute()

    rows = data.get('values', [])
    
    # update checkboxes based on UFP
    for i, row in enumerate(rows):
        used_for_prod_value = row[used_for_prod_col - 1] if len(row) > used_for_prod_col - 1 else None
        move_checkbox_value = True if used_for_prod_value else None
        
        
        requests.append({
            'updateCells': {
                'range': {
                    'sheetId': sheet_id,
                    'startRowIndex': i + 1,  # Skipping header
                    'endRowIndex': i + 2,
                    'startColumnIndex': move_col - 1,
                    'endColumnIndex': move_col
                },
                'rows': [{
                    'values': [{
                        'userEnteredValue': None
                    }]
                }],
                'fields': 'userEnteredValue'
            }
        })

    # send batch update request
    if requests:
        try:
            response = service.spreadsheets().batchUpdate(
                spreadsheetId=SPREADSHEET_ID,
                body={'requests': requests}
            ).execute()
            print(f"Successfully applied batch update for sheet {sheet_name}")
        except HttpError as err:
            print(f"HttpError: {err}")
            return

    
    print(f"Checkboxes synced in the 'Move' column for {sheet_name}")



# calculating totals
def calculate_totals(master_data):

    # calculate total on-hand sum
    total_on_hand = master_data.groupby('part_number')['on_hand'].sum().reset_index()
    total_on_hand.columns = ['part_number', 'total_on_hand']

    # total on hand cases
    total_cases_on_hand = master_data.groupby('part_number')['cases_on_hand'].sum().reset_index()
    total_cases_on_hand.columns = ['part_number', 'total_cases_on_hand']

    # merge total_on_hand DF into master_data
    master_data = master_data.merge(total_on_hand, on='part_number', how='left')
    master_data = master_data.merge(total_cases_on_hand, on='part_number', how='left')

    return master_data

# final dataframe; full proccessing of the entire data
def process_data(pick_ticket_file, inventory_file, ab_or_gf):

    # load data
    pick_data, inventory_data = load_data(pick_ticket_file, inventory_file)

    # row calculations for cheecks
    initial_pick_rows = len(pick_data)
    initial_inventory_rows = len(inventory_data)
    print(f"Initial pick rows: {initial_pick_rows}")
    print(f"Initial inventory rows: {initial_inventory_rows}")

    # process pick and inventory data
    pick_data = process_pick_data(pick_data, ab_or_gf)
    inventory_sheet = process_inventory_data(inventory_data, ab_or_gf)

    # check row counts after processing pick ticket data
    check_row_count(initial_pick_rows, len(pick_data), f"{ab_or_gf} Pick Processing")
    check_row_count(initial_inventory_rows, len(inventory_sheet), f"{ab_or_gf} Inventory Processing")

    # merge the pick and inventory data
    master_data = merge_data(pick_data, inventory_sheet)

    # format and clean the master data
    master_data = format_and_clean_master_data(master_data)

    # calculate totals
    master_data = calculate_totals(master_data)

    # final dataframe formatting
    master_data = master_data[['mo_number', 'beginning_location', 'lot_number', 'expiration_date', 'part_number', 'description',
                               'pick_qty', 'uom', 'vendor_uom', 'cases_needed', 'cases_rounded', 'staged_for_production',
                               'used_for_production', 'qty_returned_to_inventory', 'move', 'part_count', 'short_quantity',
                               'substitution', 'on_hand', 'cases_on_hand', 'total_on_hand', 'total_cases_on_hand']]

    return master_data

#INITIAL SET UP FOR SCRIPT
# (1) Open File Explorer on you laptop
# (2) In the navigation of your left, click on Documents
# (3) In the navigation at the top select "New" and create a New Folder titled: "InventoryMove"
# (4) Download the needed files, and move them to the InventoryMove Folder
      # The master data should be an excel file. The Pick Tickets and InvQtys are csv files.
# (5) Renaming
    # (1) Rename the Atom Banana Pick Ticket File to: MO Pick Ticket - Atom Banana
    # (2) Rename the Get Fresh Pick Ticket File to: MO Pick Ticket - Get Fresh
    # (3) Rename the Inventory Quantities file to: InvQtys
    # (4) Rename the Master Data Spreadsheet file to: ab_master_data
    # (5) Make sure you have the "tokens.json" and "credentials.json" in the InventoryMove Folder

# RUNNING THE SCRIPT EACH TIME
# (6) Run the script. It will ask for 2 inputs.
    # (1) Enter your name (this should be the name you use for your laptop (ex. John Smith)
    # (2) Enter the data you are looking for (ex. Atom Banana or Get Fresh)
# (7) Everything should now populate on the MK Google Sheet

        

# Ask user for their name
username = input("Enter your name: ").strip()

# Define base directory where files must be stored
base_directory = f"C:\\Users\\{username}\\Documents\\InventoryMove"

# Ensure the base directory exists
os.makedirs(base_directory, exist_ok=True)

# Ask user for Atom Banana or Get Fresh
ab_or_gf = input("Enter the data name (Atom Banana or Get Fresh): ").strip().title()

# Determine the correct pick ticket file name
if ab_or_gf == "Atom Banana":
    pick_file_name = "MO Pick Ticket - Atom Banana.csv"
elif ab_or_gf == "Get Fresh":
    pick_file_name = "MO Pick Ticket - Get Fresh.csv"
else:
    raise ValueError("Invalid input! Please enter 'Atom Banana' or 'Get Fresh'.")

# Construct full file paths
pick_file = os.path.join(base_directory, pick_file_name)
inventory_file = os.path.join(base_directory, "InvQtys.csv")
output_path = os.path.join(base_directory, "ab_master_data.xlsx")
destination_folder = os.path.join(base_directory, "InventoryMove")
shutil.copy(output_path, destination_folder)

print(f"Using pick ticket file: {pick_file}")
print(f"Using inventory file: {inventory_file}")
print(f"Processed data will be saved to: {output_path}")

# Process the data
result = process_data(pick_file, inventory_file, ab_or_gf)

# Save the new ab master data file
result.to_excel(output_path, index=False)
print(f"Processed data saved to: {output_path}")


# In[7]:


app = Flask(__name__)

#this is how the script accesses the google sheet 
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SPREADSHEET_ID = '1NuVTf2n7DGi3DhJLGz6E_Tt2GJiGt0BxqoldptK4CK8'

creds = None
token_directory = base_directory
token_filename = "token.json"
token_path = os.path.join(token_directory, token_filename)

# Create the directory if it doesn't exist
os.makedirs(token_directory, exist_ok=True)

if os.path.exists(token_path):
    creds = Credentials.from_authorized_user_file(token_path)

if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        credentials_path = os.path.join(base_directory, "credentials.json")
        flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
        creds = flow.run_local_server(port=0)

    with open(token_path, 'w') as token:
        token.write(creds.to_json())

service = googleapiclient.discovery.build('sheets', 'v4', credentials=creds)

# SCRIPT_ID = "1jtBtZXioOiPe_8OazokexBfg7CDZydtJ-FKXi2-xx-2dwQEgT8FJTP-o" 
# ACCESS_TOKEN = creds.token
# APPS_SCRIPT_URL = "https://script.google.com/a/macros/thepurplecarrot.com/s/AKfycbzrK7bv5MU66uSe0g7-QYBfX5-97KDwxPzxQ3fOmKTkMscAf_xZip2M3voKj293cjRN/exec"

# def trigger_updatePartCounts():
#     response = requests.post(APPS_SCRIPT_URL, json={"function": "updatePartCounts"})

#     if response.status_code == 200:
#         print("updatePartCounts triggered successfully!")
#     else:
#         print(f"Failed to trigger updatePartCounts. Status code: {response.status_code}, Response: {response.text}")

#     print(response.text)

def create_new_sheet(sheet_title):
    request = {
        'requests': [
            {
                'addSheet': {
                    'properties': {
                        'title': sheet_title
                    }
                }
            }
        ]
    }

    response = service.spreadsheets().batchUpdate(
    spreadsheetId=SPREADSHEET_ID,
    body=request
    ).execute()

    new_sheet_id = response['replies'][0]['addSheet']['properties']['sheetId']

    # Set the column widths for the new mk tab
    column_widths = [115,214,83,389, 100, 100,127,100,136,136,160,160,100,100,100,100,100,100,100,110, 110, 110]  # Adjust widths as needed

    for i, width in enumerate(column_widths):
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={'requests': [{'updateDimensionProperties': {'range': {'sheetId': new_sheet_id, 'dimension': 'COLUMNS', 'startIndex': i, 'endIndex': i + 1}, 'properties': {'pixelSize': width}, 'fields': 'pixelSize'}}]}
        ).execute()

    print(response)


def sheet_exists(sheet_name):
    try:
        existing_sheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID, includeGridData=False).execute()
        sheets = existing_sheet['sheets']
        sheet_titles = [sheet['properties']['title'] for sheet in sheets]
        
        return sheet_name in sheet_titles
    except Exception as e:
        print(f"Error checking sheet existence: {e}")
        return False


def copy_data_to_mk_tab(monum_value):
    source_range = f'master_data!A:Z'  # Adjust the range to cover all columns
    destination_range = f'{monum_value}!A2:Z'  # Adjust the destination range

    # Get all values from the source range
    source_values = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=source_range
    ).execute()['values']

    if not source_values:
        print(f"No data found in source range: {source_range}")
        return

    source_values = [[col for i, col in enumerate(row) if i not in [2, 3]] for row in source_values]

    expected_columns = 22
    for row in source_values:
         while len(row) < expected_columns:  # pad missing columns with None
             row.append(None)
         if len(row) > expected_columns:  # trim excess columns
             row = row[:expected_columns]

    # Create a DataFrame from the source data
    df_master_data = pd.DataFrame(source_values[1:], columns=source_values[0])

    # Filter based on both 'part_number' and 'MONUM'
    df_filtered = df_master_data[(df_master_data['part_number'].notnull()) & (df_master_data['mo_number'] == monum_value)]

    # Drop duplicates based on 'part_number' column
    df_filtered = df_filtered.drop_duplicates('part_number', keep='first').reset_index(drop=True)
    df_filtered = df_filtered.iloc[:, [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]]

    # Convert the filtered DataFrame back to a list of lists
    filtered_rows = df_filtered.values.tolist()

    # Skip copying if there are no matching rows
    if not filtered_rows:
        print(f'No matching data for sheet {monum_value}')
        return

    # Copy the filtered rows to the destination range
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=destination_range,
        valueInputOption='RAW',
        body={'values': filtered_rows}
    ).execute()

    # Add checkboxes to 'Move' and sync based on 'Used for Production'
    addCheckboxesAndSyncMove(monum_value) 




#creating the headers in the master data tab
def sheet_has_headers(sheet_name='master_data'):
    sheet = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f'{sheet_name}!A1:Z1',  # Assuming your data starts from column A
    ).execute()

    values = sheet.get('values', [])

    return len(values) > 0

# uploading the master excel sheet to the google sheet
def upload_xlsx_to_sheets():
    xlsx_file_path = f"C:\\Users\\{username}\\Documents\\InventoryMove\\ab_master_data.xlsx"
    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active

    values = []
    for row in sheet.iter_rows(values_only=True):
        values.append(list(row))

    header_row = values[0]
    data_rows = values[1:]

    # Add headers to the 'master_data' sheet if it doesn't have headers already
    if not sheet_has_headers():
        headers_range = 'master_data!A1:Z1'  # Assuming your data starts from column A
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=headers_range,
            valueInputOption='RAW',
            body={'values': [header_row]}
        ).execute()

    # Append the existing data to the 'master_data' sheet
    service.spreadsheets().values().append(
        spreadsheetId=SPREADSHEET_ID,
        range='master_data!A1',
        valueInputOption='RAW',
        body={'values': data_rows},
        insertDataOption='INSERT_ROWS'
    ).execute()

    # Create a DataFrame from the existing data
    df = pd.read_excel(xlsx_file_path)

    unique_monum_values = df['mo_number'].unique()

    for monum_value in unique_monum_values:
        print(f"Processing mo_number: {monum_value}")

        if not sheet_exists(monum_value):
            create_new_sheet(monum_value)

            # Get the new sheet range
            new_sheet_range = f'{monum_value}!A1:Z1' 

            # Add headers to the new sheet if it doesn't have headers already
            if not sheet_has_headers(monum_value):
                headers = [
                    'MONUM', 'Beginning Location', 'Part Number', 'Part Description', 'Pick Qty', 'UOM', 'Vendor UOM', 
                    'Cases Needed', 'Total Cases Rounded','Staged for Production', 'Used for Production', 
                    'Qty Returned to Inventory', 'Move','Part Count', 'short_quantity', 'Substitution',
                    'On Hand Qty', 'On Hand Cases', 'Total On Hand Qty', 'Total Cases'
                ]


                # Convert the headers to values
                headers_values = [headers]

                # Append the headers to the new sheet
                service.spreadsheets().values().update(
                    spreadsheetId=SPREADSHEET_ID,
                    range=new_sheet_range,
                    valueInputOption='RAW',
                    body={'values': headers_values}
                ).execute()
            
            copy_data_to_mk_tab(monum_value) 
            print(f'Data copied for sheet {monum_value}')
            
    ### trigger_updatePartCounts()
    print("After copying data for all sheets")

if __name__ == '__main__':
    upload_xlsx_to_sheets()


# In[ ]:




